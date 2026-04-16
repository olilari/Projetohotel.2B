import os
from flask import Flask, send_from_directory, request, jsonify
import openpyxl # Para ver e editar arquivos .xlsx
from datetime import (
    datetime,
)
# Caminho base do projeto (uma pastacima do backend)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

# Pasta frontend (HTML, JS)
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

# Pasta static (CSS)
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Pasta db (Banco de Dados)
DB_DIR = os.path.join(BASE_DIR, "db")
EXCEL_FILE = os.path.join(DB_DIR, "clientes.xlsx")

# Pasta do JS
JS_DIR = os.path.join(BASE_DIR, "frontend", "js")
JS_FILE = os.path.join(JS_DIR, "main.js")

# Cabecalhos das colunas do Excel (linha 1)
COLUMNS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro"
]

def init_excel():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR) # Cria a pasta se não existir
    
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook() # Cria a planilha
        sheet = workbook.active # Pega a planilha ativa
        sheet.title = "Clientes" # Nomeia a aba principal
        sheet.append(COLUMNS) # Adiciona os títulos das colunas
        workbook.save(EXCEL_FILE) # Salva o arquivo Excel

app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/" + STATIC_DIR)

# Pagina principal
@app.route("/")
def home():
    # return "Bom dia galera 2B!"
    return send_from_directory(FRONTEND_DIR, "index.html")

# Pagina de consulta
@app.route("/consulta")
def consulta_page():
    return send_from_directory(FRONTEND_DIR, "consulta.html")

# Pagina de alteracao
@app.route("/alterar")
def alterar_page():
    return send_from_directory(FRONTEND_DIR, "alterar.html")

# Rota para servir imagens, sripts ou outros arquivos na pasta "assets"
@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)

# -------------------------------------------------------------------------------------------------
# CADASTRAR CLIENTE
# -------------------------------------------------------------------------------------------------
@app.route("/cadastrar", methods=["POST"])
def cadastrar_cliente():
    """
    Recebe os dados do formulário (em JSON), valida e salva um novo cliente
    """
    try:
        data = request.json # Dados enviados do frontend via POST (JSON)

        # Capmos obrigatórios que o usuário deve preencher
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos."
                    }
                ),
                400,
            )
        workbook = openpyxl.load_workbook(EXCEL_FILE) # Abre o arquivo Excel
        sheet = workbook.active

        # Cria um ID automático (último ID + 1)
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        # Cria uma nova linha com os dados informados
        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacoes", ""), # Campo opcional
            datetime.now().strftime("%Y-%m-%d") # Data atual
        ]

        sheet.append(novo_cliente) # Adiciona nova linha no Excel
        workbook.save(EXCEL_FILE) # Salva alterações

        # Retorna mensagem de sucesso
        return (
            jsonify(
                {
                "status": "sucess",
                "message": "Cliente cadastrado com sucesso!",
                "id": new_id
                }
            ),
            201,
        )
    except Exception as e:
        # Tratamento deerro genérico
        return (
            jsonify ({"status": "error", "message": f"Erro ao salvar no servidor: {e}"}),
            500,
        )

@app.route("/buscar", methods={"GET"})
def buscar_clientes():

    nome_query = request.args.get("nome", "").lower()


    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet =  workbook.active
        resultados = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = dict(zip(COLUMNS,row))
            nome_cliente = (cliente.get("Nome") or ""). lower()

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados)
    
    except FileNotFoundError:
        return (
            jsonify({"status": "error", "message": "Arquivo de dados não encontrado."}),
            404,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao ler os dados: {e}"}),
            500,
        )
    

@app.route("/cliente/<int:cliente_id>", methods=["GET"])
def get_cliente(cliente_id):


    try: 
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        #procura do cliente linha por linha 
        for row_idx in range(2, sheet.max_row + 1):
            row_idx = sheet.cell(row=row_idx, column=1).value
            if row_idx == cliente_id:
                row_values = [cell.value for cell in sheet[row_idx]]
                cliente = dict(zip(COLUMNS, row_values))
                return jsonify(cliente)
            
            #Se não encontrar o ID
            return jsonify({"status": "error", "message": "Cliente não encontrado."}), 404
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao buscar cliente: {e}"}),
            500,
        )
 
@app.route("/api/atualizar/int:cliente_id>", methods=["POST"])
def atualizar_cliente(cliente_id):


    try:
        data = request.json
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active


        row_to_update = -1


        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == cliente_id:
                row_to_update = row_idx
                break

            if row_to_update == -1:
                return(
                    jsonify({
                        "status": "error",
                        "message": "Cliente não encontrado para atualização", 

                    }),
                    404,
                )

        sheet.cell(row=row_to_update, column=2, value=data.get("nome"))
        sheet.cell(row=row_to_update, column=3, value=data.get("cpf"))
        sheet.cell(row=row_to_update, column=4, value=data.get("email"))  
        sheet.cell(row=row_to_update, column=5, value=data.get("telefone"))
        sheet.cell(row=row_to_update, column=6, value=data.get("endereço"))
        sheet.cell(row=row_to_update, column=7, value=data.get("observações"))

        workbook.save(EXCEL_FILE)

        return jsonify({
            "status": "success",
            "message": "Dados do cliente atuluzados com sucesso!",
        })
    
    except Exception as e:


        return (
            jsonify({"status": "error", "message": f"Error ao atulizar dados {e}"}),
            500,
        )


if __name__ == "__main__":
    print("Base: ", BASE_DIR)
    print("Front: ", FRONTEND_DIR)
    print("Static:", STATIC_DIR)
    print("DB:", DB_DIR)
    print("Excel:", EXCEL_FILE)
    print("Pasta JS:", JS_DIR)
    print("JS:", JS_FILE)
    init_excel()
    app.run(debug=True)